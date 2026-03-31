"""Parse accrual / leave bank Excel exports (employee id + balance columns)."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def _cell_str(cell: Any) -> str:
    if cell is None or (hasattr(cell, "value") and cell.value is None):
        return ""
    v = cell.value if hasattr(cell, "value") else cell
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, bool):
        return ""
    if isinstance(v, (int, float)):
        return str(v).strip()
    return str(v).strip()


def _parse_employee_id_cell(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        if abs(value - round(value)) < 1e-9:
            return str(int(round(value)))
        return None
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    if re.fullmatch(r"-?\d+", s):
        return str(int(s))
    return None


def _format_balance(value: Any) -> str:
    if value is None or value == "":
        return "0.00"
    try:
        if isinstance(value, str):
            x = float(value.replace(",", "").strip())
        else:
            x = float(value)
    except (ValueError, TypeError):
        return "0.00"
    return f"{x:.2f}"


def parse_leavebank_workbook(
    content: bytes,
) -> tuple[dict[str, list[dict[str, str]]], list[str]]:
    """
    Returns (banks_by_employee_id, warnings).

    Each value is a list of {"key", "label", "value"} for display, in column order.
    """
    warnings: list[str] = []
    banks: dict[str, list[dict[str, str]]] = {}

    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as e:
        warnings.append(f"Leave bank: could not open workbook ({e}).")
        return {}, warnings

    try:
        ws = wb.active
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row < 2 or max_col < 2:
            warnings.append("Leave bank: sheet is empty or too small.")
            return {}, warnings

        header_row_idx: int | None = None
        for r in range(1, min(21, max_row + 1)):
            a = _cell_str(ws.cell(row=r, column=1))
            if a.lower() == "employee":
                header_row_idx = r
                break

        if header_row_idx is None:
            warnings.append('Leave bank: could not find a header row with "Employee" in column A.')
            return {}, warnings

        specs: list[tuple[int, str, str]] = []
        used_keys: set[str] = set()
        for c in range(2, max_col + 1):
            label = _cell_str(ws.cell(row=header_row_idx, column=c))
            if not label:
                continue
            if label.lower() in ("name", "employee name"):
                continue
            key = label.split(" - ", 1)[0].strip() if " - " in label else label
            base_key = key
            k = key
            n = 1
            while k in used_keys:
                n += 1
                k = f"{base_key}_{n}"
            used_keys.add(k)
            specs.append((c, label, k))

        if not specs:
            warnings.append("Leave bank: no balance columns found after the header row.")
            return {}, warnings

        seen_emp: set[str] = set()
        for r in range(header_row_idx + 1, max_row + 1):
            eid = _parse_employee_id_cell(ws.cell(row=r, column=1).value)
            if eid is None:
                continue
            if eid in seen_emp:
                warnings.append(f"Leave bank: duplicate employee id {eid}; using last row.")
            seen_emp.add(eid)
            cols: list[dict[str, str]] = []
            for c, label, key in specs:
                v = ws.cell(row=r, column=c).value
                cols.append(
                    {
                        "key": key,
                        "label": label,
                        "value": _format_balance(v),
                    }
                )
            banks[eid] = cols

    finally:
        wb.close()

    return banks, warnings
